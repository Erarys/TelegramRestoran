from datetime import datetime, timedelta

from aiogram import Router, F
from aiogram.filters import Command, CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, ReplyKeyboardMarkup, CallbackQuery
from aiogram.types.input_file import FSInputFile

from db.queries.check_get import get_menu
from db.queries.orm import create_report, create_report_period, fill_table, fill_food_menu, delete_menu, \
    create_food_report, restart_table, delete_orders
from filters.base_filters import IsAdmin
from keyboards.admin_keyboard import get_menu_button, FoodDeleteCallback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class MenuForm(StatesGroup):
    name: str = State()
    price: int = State()


router = Router()
router.message.filter(IsAdmin())
filter_for_lagman = [
    "Гуйру",
    "Суйру",
    "Домашний лагман",
    "Гуйру цомян",
    "Дин-дин",
    "Лагман с ребрами",
    "Могру",
    "Хаухуа",
    "Мошру",
    "Фирменный лагман \"Арыс\"",
    "Красные пельмени",
    "Мампар",
    "Суп с мясом",
    "Пельмень",
    "Фри с мясом",
    "Мясо по-тайски",
    "Мушуру сай",
    "Могуру сай",
    "Казан-кебаб",
    "Дапанджи",
    "Свежий салат",
    "Пекинский салат",
    "Хрустящий баклажан",
    "Ачучук",
    "Фри",
]
filter_for_shashlik = [
    "Баранина",
    "Утка",
    "Крылочка",
    "Люля",
    "Ребра",
    "Антрекот",
    "Говядина",
    "Овощной",
    "Шампинион",
    "Окорачка",
]

def excel_work(orders_df, excel_path):
    # Предположим, что это DataFrame
    df = orders_df

    # Сохраняем во временный файл
    df.to_excel(excel_path, index=False)

    # Загружаем книгу и лист
    wb = load_workbook(excel_path)
    ws = wb.active

    # 🎨 Оформление заголовка
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # светло-жёлтый
    header_font = Font(bold=True)

    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font

        # 📏 Автоширина
        max_length = len(str(cell.value))
        col_letter = get_column_letter(col_num)

        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell_in_row in row:
                if cell_in_row.value:
                    max_length = max(max_length, len(str(cell_in_row.value)))

        ws.column_dimensions[col_letter].width = max_length + 2  # небольшой отступ

    # 🎨 Заливка для последней строки
    last_row_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # светло-зелёный

    for cell in ws[ws.max_row]:  # последняя строка
        cell.fill = last_row_fill

    # Сохраняем обновлённый файл
    wb.save(excel_path)


@router.message(Command("delete"))
async def delete_order(message: Message, command: CommandObject):
    await message.answer(f"{message.chat.id}")
    if command.args is not None:
        figure = command.args.split()
        result = await delete_orders(list(map(int, figure)))
        await message.answer(f"Мы удалили: {result}")


@router.message(Command("report"))
async def create_order_report(message: Message, command: CommandObject):
    await message.answer(f"{message.chat.id}")
    if command.args is not None:
        try:
            start_day, end_day = command.args.split()
            file_path = f"reports/report_{start_day} {end_day}.xlsx"
            start_day_time = datetime.strptime(start_day, "%d.%m.%Y")
            end_day_time = datetime.strptime(end_day, "%d.%m.%Y")

            orders_dt = await create_report_period(start_day_time, end_day_time)
        except:
            await message.answer("Ошибка")
    else:
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)
        file_path = f"reports/report_{today:%Y%m%d}.xlsx"

        orders_dt: dict = await create_report(today, tomorrow)

    orders_dt["Итого"] = {
        "Чек": sum([value["Чек"] for value in orders_dt.values()])
    }
    orders_df = pd.DataFrame.from_dict(orders_dt, orient='index')  # ключи 10, 11 станут индексами
    excel_work(orders_df, file_path)

    document = FSInputFile(file_path)
    await message.bot.send_document(message.chat.id, document)


@router.message(Command("report_food"))
async def report_shashlik(message: Message, command: CommandObject):
    args = command.args
    if not isinstance(args, str):
        await message.answer("Вы не передали ни одного аргумента")

    if args is not None:
        try:
            start_day, end_day = command.args.split()
            file_path = f"reports/report_{start_day} {end_day}.xlsx"
            start_day_time = datetime.strptime(start_day, "%d.%m.%Y")
            end_day_time = datetime.strptime(end_day, "%d.%m.%Y")

            orders_dt = await create_food_report(start_day_time, end_day_time, filter_for_shashlik)
        except:
            await message.answer("Ошибка")
            return
    else:
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)

        file_path = f"reports/report_food{today:%Y%m%d}.xlsx"

        orders_dt: dict = await create_food_report(today, tomorrow, filter_for_shashlik)

    orders_dt["Итого"] = {
        "Сумма": sum([value["Сумма"] for value in orders_dt.values()]),
        "Доля шашлычника": sum([value["Доля шашлычника"] for value in orders_dt.values()]),
    }

    orders_df = pd.DataFrame.from_dict(orders_dt, orient='index')  # ключи 10, 11 станут индексами
    excel_work(orders_df, file_path)

    document = FSInputFile(file_path)
    await message.bot.send_document(message.chat.id, document)


@router.message(Command("report_food2"))
async def report_lagman(message: Message, command: CommandObject):
    args = command.args
    if not isinstance(args, str):
        await message.answer("Вы не передали ни одного аргумента")

    if args is not None:
        try:
            start_day, end_day = command.args.split()
            file_path = f"reports/report_{start_day} {end_day}.xlsx"
            start_day_time = datetime.strptime(start_day, "%d.%m.%Y")
            end_day_time = datetime.strptime(end_day, "%d.%m.%Y")

            orders_dt = await create_food_report(start_day_time, end_day_time, filter_for_lagman)
        except:
            await message.answer("Ошибка")
            return
    else:
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)
        file_path = f"reports/report_food{today:%Y%m%d}.xlsx"

        orders_dt: dict = await create_food_report(today, tomorrow, food_names=filter_for_lagman)

    orders_dt["Итого"] = {
        "Сумма": sum([value["Сумма"] for value in orders_dt.values()]),
        "Доля шашлычника": sum([value["Доля шашлычника"] for value in orders_dt.values()]),
    }

    orders_df = pd.DataFrame.from_dict(orders_dt, orient='index')  # ключи 10, 11 станут индексами
    excel_work(orders_df, file_path)

    document = FSInputFile(file_path)
    await message.bot.send_document(message.chat.id, document)

@router.message(Command("add_table"))
async def restart_order(message: Message, command: CommandObject):
    args = command.args
    if not isinstance(args, str):
        await message.answer("Вы не передали ни одного аргумента")

    if args.isdigit():
        amount = int(args)
        await fill_table(amount)
        await message.answer("Таблица создана")
    else:
        await message.answer("Аргумент должно быть целым числом!")


@router.callback_query(FoodDeleteCallback.filter())
async def delete_food(callback: CallbackQuery, callback_data: FoodDeleteCallback):
    food_id = callback_data.food_id
    await delete_menu(food_id)
    await callback.message.edit_text("Удалено", reply_markup=None)


@router.message(Command("restart"))
async def restart_table_handler(message: Message):
    await restart_table()
    await message.answer("Все столы были сброшены 🚮")


@router.message(Command("update_menu"))
async def update_menu(message: Message, state: FSMContext):
    await state.clear()
    menu: dict = await get_menu(0, 10000)

    for id in menu.keys():
        await message.answer(
            text=f"{menu[id]['name']}\n"
                 f"{menu[id]['price']}\n"
                 f"Введите название продукта:",
            reply_markup=get_menu_button(id)
        )

    await message.answer(text="Введите название продукта:")
    await state.set_state(MenuForm.name)


@router.message(F.text, MenuForm.name)
async def food_name(message: Message, state: FSMContext):
    name = message.text
    await state.update_data(name=name)

    await message.answer(text="Введите цену:")
    await state.set_state(MenuForm.price)


@router.message(F.text, MenuForm.price)
async def food_price(message: Message, state: FSMContext):
    price = message.text
    await state.update_data(price=price)
    food = await state.get_data()

    if food.get("name") and food.get("price"):
        is_exist = await fill_food_menu(food.get("name"), food.get("price"))
        print(food.get("name"))
        if is_exist:
            await message.answer("Продукт уже существует, сначала удалите старый 🚨")
        else:
            await message.answer("Продукт успешно добавлен")
    else:
        await message.answer("Ошибка")


@router.message(Command("auto_fill"))
async def auto_fill(message: Message):
    dt = {
        "Гуйру": 2200,
        "Суйру": 2200,
        "Домашний лагман": 2200,
        "Гуйру цомян": 2200,
        "Дин-дин": 2200,
        "Лагман с ребрами": 2200,
        "Могру": 2200,
        "Хаухуа": 2200,
        "Мошру": 2500,
        "Фирменный лагман \"Арыс\"": 2500,
        "Красные пельмени": 2000,
        "Мампар": 2000,
        "Суп с мясом": 2000,
        "Пельмень": 2000,
        "Фри с мясом": 2800,
        "Пюре": 1000,
        "Свежий салат": 1500,
        "Пекинский салат": 2500,
        "Хрустящий баклажан": 2500,
        "Ачучук": 2000,
        "Coca-Cola 2л": 1000,
        "Coca-Cola 1.5л": 800,
        "Fanta 1л": 700,
        "Coca-Cola 1л": 700,
        "Чай": 350,
        "Фри": 1000,
        "Спиральные чипсы": 700,
        "Лепешка": 250,
        "Детский сок": 250,

        "Прага": 800,
        "Гусь Крепкое": 900,
        "Гусь Экспортное": 800,
        "Гусь": 800,
        "Carlsberg": 800,
        "Holsten Pilsener":900,
        "Holsten Light": 900,
        "Garage Lemon": 900,
        "Garage Cherry": 900,
        "Добрый Бобр": 800,
        "Добрый Крепкое": 800,
        "Yang": 800,
        "Flash": 600,
        "Водка 100мл": 700,



        "Баранина":2000,
        "Утка": 1600,
        "Крылочка": 1500,
        "Люля": 1800,
        "Ребра": 1800,
        "Антрекот": 2400,
        "Говядина": 2000,
        "Овощной": 1000,
        "Шампинион":1300,
        "Курт": 200,
        "Turan Вода": 450,
        "Лейс лук": 1000,
        "Лейс сыр": 700,
        "Фан": 700,
        "Кириешки": 350,
        "Багет": 350,
        "Кириешки лайт": 150,
        "Арахис": 500,
        "Арахис лайт": 400,
        "Окорачка": 1800,


    }

    dt2 = {
        "Мясо по-тайски": 3000,
        "Мушуру сай": 3000,
        "Могуру сай": 3000,
        "Казан-кебаб": 3000,
        "Дапанджи": 10000,
    }

    garnishes = ["рис", "фри", "пюре"]

    # Новый словарь с блюдами и гарнирами
    menu_with_garnish = {}

    for dish, price in dt2.items():
        for garnish in garnishes:
            dish_with_garnish = f"{dish} ({garnish})"
            menu_with_garnish[dish_with_garnish] = price

    dt.update(menu_with_garnish)
    for food_name, price in dt.items():
        is_exist = await fill_food_menu(food_name, price)
        if is_exist:
            await message.answer(f"Обновили цену: {food_name}: {price}")
        else:
            await message.answer(f"{food_name}: {price}")
